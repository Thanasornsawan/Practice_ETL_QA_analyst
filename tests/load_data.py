import sqlite3
import os
from openpyxl import load_workbook
from datetime import datetime

# Path to the SQLite database
DB_PATH = 'sql/sqlite_db/etl.db'
# Path to the Excel file (dynamically resolve the absolute path)
EXCEL_FILE_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '../orders_test_data.xlsx'))

def load_data_to_db():
    # Load the workbook and the 'Products' and 'Orders' sheets
    wb = load_workbook(EXCEL_FILE_PATH)

    # Access the 'Products' and 'Orders' sheets
    products_sheet = wb['Products']
    orders_sheet = wb['Orders']

    # Establish a database connection
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Insert data into Products table
    for row in products_sheet.iter_rows(min_row=2, values_only=True):
        cursor.execute('''
            INSERT OR IGNORE INTO Products (Product_ID, Product_Name) 
            VALUES (?, ?)
        ''', (row[0], row[1]))

    # Insert data into Orders table
    for row in orders_sheet.iter_rows(min_row=2, values_only=True):
        # Check if the row is empty (all fields are empty or None)
        if all(cell is None or cell == '' for cell in row):
            continue  # Skip the row if it's empty

        customer_id = row[0]
        customer_name = row[1]
        order_date = row[2]  # Order_Date is assumed to be in the third column (index 2)
        product_id = row[3]
        quantity = row[4]
        email = row[5]

        # Ensure that 'Order_Date' stays as a string, not a date object
        if isinstance(order_date, str):
            # If the order_date is in string format (like '12/01/2024'), keep it as is
            order_date = order_date.strip()  # Remove leading/trailing whitespace and newlines
        elif isinstance(order_date, datetime):
            # If the order_date is a datetime object, convert it to string
            order_date = order_date.strftime('%d/%m/%Y') if order_date else None
        else:
            order_date = None  # Set to None if the date format is invalid

        # Skip inserting rows where required data (such as order_date or customer_id) is invalid
        if not customer_id or not order_date:
            continue  # Skip this row if customer_id or order_date is missing or invalid

        cursor.execute('''
            INSERT INTO Orders (Customer_ID, Customer_Name, Order_Date, Product_ID, Quantity, Email) 
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (customer_id, customer_name, order_date, product_id, quantity, email))

    # Commit the changes and close the connection
    conn.commit()
    conn.close()

    print("Data loaded successfully from Excel to database.")

if __name__ == '__main__':
    load_data_to_db()
